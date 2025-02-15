import os
import pandas as pd
import google.generativeai as genai
from openpyxl import load_workbook

# Configure API Key
api = "Enter Your API key"  # Ensure your API key is valid
genai.configure(api_key=api)

# Function to upload a file
def upload_file(path, mime_type=None):
    file = genai.upload_file(path, mime_type=mime_type)
    return file

# Function to process images and save results
def process_images(folder_path, output_file, start_row=2):  # Default start from row 2 (row 1 for headers)
    generation_config = {
        "temperature": 1,
        "top_p": 0.95,
        "top_k": 40,
        "max_output_tokens": 8192,
        "response_mime_type": "text/plain",
    }

    model = genai.GenerativeModel(
        model_name="gemini-2.0-flash",
        generation_config=generation_config,
    )

    # Supported image formats
    valid_extensions = {".jpg", ".jpeg", ".png", ".gif"}

    # List to store results
    results = []

    # Iterate through all files in the folder
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)

        # Check if the file is an image
        if os.path.isfile(file_path) and os.path.splitext(file_name)[1].lower() in valid_extensions:
            file = upload_file(file_path, mime_type="image/png")  # Change mime_type if needed

            # Generate response
            response = model.generate_content([
                file,
                "Extract only the text from the image. No need to write any additional things. if there is no text in the image then write 'No text found'. if there is any prohibited content then write 'prohibited' "
            ])

            extracted_text = response.text.strip()  # Clean text

            # Debugging: Print output
            print(f"Processing: {file_name} ---> Extracted Text: {extracted_text}")

            # Append result to list
            results.append({"Image Name": file_name, "Image Text": extracted_text})

    # Convert results to DataFrame
    new_data = pd.DataFrame(results)

    if os.path.exists(output_file):
        # Load the existing workbook
        book = load_workbook(output_file)
        sheet = book.active

        # Get the last **actual** occupied row (ignoring empty rows)
        last_row = sheet.max_row
        while last_row > 1 and all(sheet[row][0].value is None for row in range(last_row, last_row - 3, -1)):
            last_row -= 1  # Move up if empty rows exist

        # Manually set where to append new data
        append_row = max(last_row + 1, start_row)

        # Load existing data
        existing_data = pd.read_excel(output_file)

        # Assign new serial numbers correctly
        last_sl_no = existing_data["Sl no"].max() if not existing_data.empty else 0
        new_data.insert(0, "Sl no", range(last_sl_no + 1, last_sl_no + len(new_data) + 1))

        # Append new data exactly at `append_row`
        with pd.ExcelWriter(output_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            new_data.to_excel(writer, index=False, header=False, startrow=append_row)

    else:
        # If file doesn't exist, create a new one
        new_data.insert(0, "Sl no", range(1, len(new_data) + 1))
        new_data.to_excel(output_file, index=False)



    print(f"\nResults saved to {output_file}")

# Provide the folder path here
folder_path = r"C:\Users\User\Downloads\Telegram Desktop\ChatExport_2025-02-12\photos\601-700"  # Change this to your actual folder path
output_file = r"C:\Users\User\Downloads\Telegram Desktop\ChatExport_2025-02-12\photos\Exel\5.xlsx"  # Output Excel file
start_row = 1  # Set the manual row number where data should start

process_images(folder_path, output_file, start_row)
